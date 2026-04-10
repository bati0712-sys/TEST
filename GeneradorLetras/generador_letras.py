"""
generador_letras.py — Generador de Letras de Cambio
Replica y mejora GeneradorLetras_v13.xlsm
"""
import sys, json, math
from pathlib import Path
from datetime import date, datetime, timedelta
from dateutil.relativedelta import relativedelta

try:
    import pyodbc
    _PYODBC_OK = True
except ImportError:
    _PYODBC_OK = False

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QTabWidget,
    QVBoxLayout, QHBoxLayout, QFormLayout, QGridLayout,
    QLabel, QLineEdit, QPushButton, QComboBox, QDateEdit,
    QSpinBox, QDoubleSpinBox, QTableWidget, QTableWidgetItem,
    QHeaderView, QGroupBox, QMessageBox, QDialog,
    QTextEdit, QCheckBox, QSplitter, QFrame, QScrollArea,
    QAbstractItemView, QSizePolicy,
)
from PySide6.QtCore import Qt, QDate, Signal, QThread, QSizeF
from PySide6.QtGui import QFont, QColor, QBrush, QTextDocument, QPageSize
try:
    from PySide6.QtPrintSupport import QPrinter, QPrintPreviewDialog
    _PRINT_OK = True
except ImportError:
    _PRINT_OK = False

# ── Rutas ─────────────────────────────────────────────────────────────────────
if getattr(sys, 'frozen', False):
    BASE = Path(sys.executable).parent
else:
    BASE = Path(__file__).parent

DATA_DIR = BASE / 'data'
DATA_DIR.mkdir(exist_ok=True)

F_BASE       = DATA_DIR / 'base.json'
F_PARAMS     = DATA_DIR / 'parametros.json'
F_HIST       = DATA_DIR / 'historico.json'
F_EXT        = DATA_DIR / 'externos.json'
F_DYSPOW_CFG = DATA_DIR / 'dyspow_config.json'


# ── Persistencia ──────────────────────────────────────────────────────────────
import threading
_save_lock = threading.Lock()

def _load(path: Path, default):
    if path.exists():
        for _ in range(5):   # reintentos si el archivo está bloqueado
            try:
                with open(path, encoding='utf-8') as f:
                    return json.load(f)
            except (PermissionError, OSError):
                import time; time.sleep(0.2)
            except Exception:
                break
    return default

def _save(path: Path, data):
    """Escritura atómica: escribe en .tmp y renombra, con lock global."""
    with _save_lock:
        tmp = path.with_suffix('.tmp')
        for _ in range(5):
            try:
                with open(tmp, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2, default=str)
                tmp.replace(path)
                return
            except (PermissionError, OSError):
                import time; time.sleep(0.3)
        raise IOError(f"No se pudo guardar {path.name} — archivo en uso por otro usuario.")

def load_base():
    return _load(F_BASE, {
        "empresas":   ["SHOE TRADE"],
        "proveedores": [{"nombre": "PROVEEDOR EJEMPLO SAC", "ruc": "20123456789"}],
        "temporadas": ["LT - "],
        "monedas": [
            {"codigo": "PEN", "nombre": "Sol Peruano",       "tc": 1.0},
            {"codigo": "USD", "nombre": "Dólar Americano",   "tc": 3.70},
            {"codigo": "EUR", "nombre": "Euro",              "tc": 4.00},
            {"codigo": "GBP", "nombre": "Libra Esterlina",   "tc": 4.60},
        ]
    })

def save_base(data):   _save(F_BASE, data)

def load_params():
    return _load(F_PARAMS, {
        "presupuesto_default": 50000,
        "separacion_minima":   0,
        "monto_minimo":        5000,
        "monto_maximo":        30000,
        "limite_diario_max":   0,
        "parcial_minimo":      1000,
        "parcial_maximo":      4999,
        "dias_bloqueados": {
            "Lunes": False, "Martes": False, "Miércoles": False,
            "Jueves": False, "Viernes": False, "Sábado": True, "Domingo": True
        },
        "periodos_bloqueados": [],
        "presupuesto_mensual": []
    })

def save_params(data): _save(F_PARAMS, data)

def load_hist():
    return _load(F_HIST, [])

def save_hist(data):   _save(F_HIST, data)

def load_ext():
    return _load(F_EXT, [])

def save_ext(data):    _save(F_EXT, data)

def load_dyspow_cfg():
    return _load(F_DYSPOW_CFG, {
        "server": "192.168.1.114",
        "database": "BD_DYSPOW_BFX",
        "user": "DyspowUser",
        "password": "Dyspow",
    })

def save_dyspow_cfg(data): _save(F_DYSPOW_CFG, data)


# ── Motor de cálculo ──────────────────────────────────────────────────────────
def r500(v: float) -> float:
    """Redondea al múltiplo de 500 más cercano."""
    return round(round(v / 500) * 500, 2)

def es_bloqueado(f: date, params: dict) -> bool:
    dias = ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"]
    nombre_dia = dias[f.weekday()]
    if params["dias_bloqueados"].get(nombre_dia, False):
        return True
    for p in params.get("periodos_bloqueados", []):
        try:
            fi = date.fromisoformat(p["inicio"])
            ff = date.fromisoformat(p["fin"])
            if fi <= f <= ff:
                return True
        except Exception:
            pass
    return False

def fecha_valida_atras(f: date, params: dict) -> date:
    for _ in range(90):
        if not es_bloqueado(f, params):
            return f
        f -= timedelta(days=1)
    return f

def dias_habiles_mes(anio: int, mes: int, params: dict) -> int:
    d = date(anio, mes, 1)
    if mes == 12:
        fin = date(anio + 1, 1, 1) - timedelta(days=1)
    else:
        fin = date(anio, mes + 1, 1) - timedelta(days=1)
    count = 0
    while d <= fin:
        if not es_bloqueado(d, params):
            count += 1
        d += timedelta(days=1)
    return count

def get_tc(params_base: dict, moneda: str) -> float:
    if moneda.upper() == "PEN":
        return 1.0
    for m in params_base.get("monedas", []):
        if m["codigo"].upper() == moneda.upper():
            return float(m.get("tc", 1.0))
    return 1.0

def get_lim_diario(f: date, params: dict) -> float:
    for row in params.get("presupuesto_mensual", []):
        try:
            mes_date = date.fromisoformat(row["mes"] + "-01")
        except Exception:
            continue
        if mes_date.year == f.year and mes_date.month == f.month:
            ov = float(row.get("override", 0) or 0)
            if ov > 0:
                return ov
            presup = float(row.get("presupuesto", 0) or 0)
            if presup > 0:
                dh = dias_habiles_mes(f.year, f.month, params)
                if dh > 0:
                    return presup / dh
    # fallback: presupuesto_default / días hábiles del mes
    presup_def = float(params.get("presupuesto_default", 0) or 0)
    if presup_def > 0:
        dh = dias_habiles_mes(f.year, f.month, params)
        if dh > 0:
            return presup_def / dh
    return float(params.get("limite_diario_max", 0) or 0)

def comp_dia(f: date, historico: list, externos: list, params_base: dict) -> float:
    """Monto comprometido en PEN para la fecha f (histórico + externos)."""
    total = 0.0
    for h in historico:
        try:
            fv = date.fromisoformat(h["fecha_vcto"])
        except Exception:
            continue
        if fv == f:
            tc = get_tc(params_base, h.get("moneda", "PEN"))
            total += float(h.get("monto", 0)) * tc
    for e in externos:
        try:
            fv = date.fromisoformat(e["fecha_vcto"])
        except Exception:
            continue
        if fv == f:
            total += float(e.get("monto_pen", 0))
    return total

def cyc_dia(f: date, cyc: list) -> float:
    """Monto asignado en ciclo actual para la fecha f."""
    return sum(c["monto_pen"] for c in cyc if c["fecha"] == f)

def gen_id(historico: list) -> str:
    max_n = 0
    for h in historico:
        id_op = h.get("id_operacion", "")
        if id_op.startswith("OP-"):
            try:
                n = int(id_op[3:])
                if n > max_n:
                    max_n = n
            except Exception:
                pass
    return f"OP-{max_n + 1:04d}"

def next_num(historico: list, temporada: str) -> int:
    return sum(1 for h in historico if h.get("temporada", "") == temporada) + 1

def generar_cronograma(
    empresa, temporada, proveedor, ruc, concepto, observac, moneda,
    num_letras_deseadas, fecha_emi, dias_cred, fecha_lim,
    facturas, dist_manual,
    params, params_base, historico, externos
):
    """
    Retorna (letras, avisos) donde letras = lista de dicts.
    Lanza ValueError si hay validaciones críticas.
    """
    tc = get_tc(params_base, moneda)

    neto_total = sum(f["importe"] - f["retencion"] for f in facturas)
    imp_bruto  = sum(f["importe"] for f in facturas)
    ret_total  = sum(f["retencion"] for f in facturas)

    if neto_total <= 0:
        raise ValueError("El Neto Total debe ser mayor a 0.")

    neto_pen = neto_total * tc

    monto_min = float(params.get("monto_minimo", 5000) or 5000)
    monto_max = float(params.get("monto_maximo", 30000) or 30000)
    lim_dia_max = float(params.get("limite_diario_max", 0) or 0)
    parcial_min = float(params.get("parcial_minimo", 0) or 0)
    parcial_max = float(params.get("parcial_maximo", 0) or 0)
    sep_min     = int(params.get("separacion_minima", 0) or 0)

    if monto_max <= 0 or monto_min <= 0:
        raise ValueError("Los parámetros Monto Mínimo y Máximo de letra no pueden ser 0.")

    MAX_LETRAS = 50
    TOL = 0.01

    # Calcular número de letras
    if neto_pen < monto_min:
        num_calc = 1
    elif num_letras_deseadas > 0:
        num_calc = num_letras_deseadas
        test = neto_pen / num_calc
        if test > monto_max + TOL:
            sug = math.ceil(neto_pen / monto_max)
            num_calc = sug  # ajusta automáticamente
    else:
        num_calc = round(neto_pen / ((monto_min + monto_max) / 2))
        if num_calc < 1:
            num_calc = 1

    if num_calc > MAX_LETRAS:
        raise ValueError(f"Se requieren {num_calc} letras pero el máximo es {MAX_LETRAS}.")

    avisos = []
    if num_letras_deseadas > 0 and num_calc != num_letras_deseadas:
        avisos.append(f"Se solicitaron {num_letras_deseadas} letras; se generarán {num_calc} por ajuste de límites.")

    # Construir índice de distribución manual (mes → límite)
    dist_idx = {}
    for d in dist_manual:
        try:
            mes_d = date.fromisoformat(d["mes"] + "-01")
            dist_idx[(mes_d.year, mes_d.month)] = float(d["limite"])
        except Exception:
            pass

    # Algoritmo: backwards desde fecha_lim
    cyc         = []          # letras del ciclo actual
    res_fechas  = []
    res_montos  = []          # en moneda original
    res_espacios= []

    monto_rest    = neto_total
    monto_rest_pn = neto_pen
    fecha_actual  = fecha_lim
    max_iter      = 1000
    iters         = 0

    while monto_rest_pn > 0.005 and iters < max_iter and len(res_fechas) < MAX_LETRAS:
        iters += 1
        fecha_actual = fecha_valida_atras(fecha_actual, params)

        if fecha_actual < fecha_emi:
            avisos.append(f"No hay suficientes días hábiles. Letras generadas: {len(res_fechas)}")
            break

        # Comprometido total en PEN para este día
        comp_h = comp_dia(fecha_actual, historico, externos, params_base)
        comp_c = cyc_dia(fecha_actual, cyc)
        comp_total = comp_h + comp_c

        # Verificar separación mínima
        if sep_min > 0 and cyc:
            demasiado = any(
                abs((fecha_actual - c["fecha"]).days) < sep_min
                for c in cyc
            )
            if demasiado:
                fecha_actual -= timedelta(days=1)
                continue

        # Límite diario efectivo
        key = (fecha_actual.year, fecha_actual.month)
        if key in dist_idx:
            lim_efec = dist_idx[key]
        else:
            lim_efec = get_lim_diario(fecha_actual, params)

        espacio = max(lim_efec - comp_total, 0.0)

        if espacio < parcial_min:
            fecha_actual -= timedelta(days=1)
            continue

        # Monto a asignar
        is_ultima = len(res_fechas) >= num_calc - 1 or monto_rest_pn <= monto_max + TOL
        pendiente = monto_rest_pn if is_ultima else (monto_min + monto_max) / 2

        monto_let_pn = 0.0
        if espacio >= pendiente:
            monto_let_pn = r500(pendiente)
            if monto_let_pn > monto_max + TOL:
                monto_let_pn = r500(monto_max)
        elif espacio >= monto_min:
            monto_let_pn = r500(espacio)
            if monto_let_pn > monto_max + TOL:
                monto_let_pn = r500(monto_max)
        elif espacio >= parcial_min:
            monto_let_pn = r500(espacio)
            if monto_let_pn > parcial_max:
                monto_let_pn = r500(parcial_max)
        else:
            fecha_actual -= timedelta(days=1)
            continue

        if monto_let_pn > monto_rest_pn:
            monto_let_pn = round(monto_rest_pn, 2)
        if monto_let_pn <= 0:
            fecha_actual -= timedelta(days=1)
            continue

        # Absorber residual pequeño
        residual_pn = monto_rest_pn - monto_let_pn
        if 0 < residual_pn < parcial_min:
            if monto_let_pn + residual_pn <= monto_max + TOL:
                monto_let_pn = round(monto_let_pn + residual_pn, 2)
            elif res_fechas:
                total_ult = (res_montos[-1] * tc) + monto_let_pn + residual_pn
                mitad1 = r500(total_ult / 2)
                res_montos[-1] = mitad1 / tc
                monto_let_pn = round(total_ult - mitad1, 2)

        monto_let_orig = round(monto_let_pn / tc, 2) if tc > 0 else monto_let_pn

        res_fechas.append(fecha_actual)
        res_montos.append(monto_let_orig)
        res_espacios.append(espacio)

        monto_rest    -= monto_let_orig
        monto_rest_pn -= monto_let_pn
        if monto_rest    < 0.005: monto_rest    = 0
        if monto_rest_pn < 0.005: monto_rest_pn = 0

        cyc.append({"fecha": fecha_actual, "monto_pen": monto_let_pn})
        fecha_actual -= timedelta(days=1)

    if not res_fechas:
        raise ValueError("No se pudo generar ninguna letra. Verifique presupuestos y días hábiles.")

    if monto_rest_pn > 0.005:
        saldo = round(monto_rest_pn / tc, 2)
        avisos.append(f"Saldo sin asignar: {moneda} {saldo:,.2f} — revise presupuestos o divida la operación.")

    # Invertir: de cronológico hacia adelante
    res_fechas.reverse()
    res_montos.reverse()
    res_espacios.reverse()

    start_n = next_num(historico, temporada)
    letras = []
    for i, (f, m, esp) in enumerate(zip(res_fechas, res_montos, res_espacios)):
        monto_pen = round(m * tc, 2)
        if monto_pen >= parcial_min and monto_pen < monto_min:
            estado = "Parcial"
        elif monto_pen < parcial_min:
            estado = "Ajuste"
        else:
            estado = "Pendiente"
        letras.append({
            "n":           i + 1,
            "codigo":      f"{temporada}{start_n + i:04d}",
            "fecha_emi":   fecha_emi.isoformat(),
            "fecha_vcto":  f.isoformat(),
            "moneda":      moneda,
            "espacio":     round(esp, 2),
            "monto":       m,
            "monto_pen":   monto_pen,
            "estado":      estado,
            "empresa":     empresa,
            "temporada":   temporada,
            "proveedor":   proveedor,
            "ruc":         ruc,
            "concepto":    concepto,
            "observac":    observac,
        })

    return letras, avisos


# ── Helpers UI ─────────────────────────────────────────────────────────────────
def make_item(txt, align=Qt.AlignLeft | Qt.AlignVCenter, editable=False):
    item = QTableWidgetItem(str(txt))
    item.setTextAlignment(align)
    if not editable:
        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
    return item

def num_item(v, fmt="{:,.2f}", align=Qt.AlignRight | Qt.AlignVCenter):
    try:
        return make_item(fmt.format(float(v)), align)
    except Exception:
        return make_item(str(v), align)


# ── Tab: Generador ─────────────────────────────────────────────────────────────
class TabGenerador(QWidget):
    cronograma_generado = Signal(list)   # emite letras generadas

    def __init__(self, get_data, parent=None):
        super().__init__(parent)
        self._get_data = get_data  # callback → (params, params_base, hist, ext)
        self._letras   = []
        self._build()

    def _build(self):
        root = QVBoxLayout(self)
        root.setSpacing(6)

        # ── Datos de la operación ─────────────────────────────────────────────
        grp_op = QGroupBox("Datos de la Operación")
        g = QGridLayout(grp_op)

        g.addWidget(QLabel("Empresa Emisora *"), 0, 0)
        self.cmb_empresa = QComboBox(); self.cmb_empresa.setEditable(True)
        g.addWidget(self.cmb_empresa, 0, 1)

        g.addWidget(QLabel("Temporada *"), 0, 2)
        self.cmb_temp = QComboBox(); self.cmb_temp.setEditable(True)
        g.addWidget(self.cmb_temp, 0, 3)

        g.addWidget(QLabel("Proveedor *"), 1, 0)
        self.cmb_prov = QComboBox(); self.cmb_prov.setEditable(True)
        self.cmb_prov.currentTextChanged.connect(self._on_prov_change)
        g.addWidget(self.cmb_prov, 1, 1)

        g.addWidget(QLabel("RUC"), 1, 2)
        self.txt_ruc = QLineEdit(); self.txt_ruc.setMaximumWidth(160)
        g.addWidget(self.txt_ruc, 1, 3)

        g.addWidget(QLabel("Concepto *"), 2, 0)
        self.txt_concepto = QLineEdit()
        g.addWidget(self.txt_concepto, 2, 1, 1, 3)

        g.addWidget(QLabel("Observaciones"), 3, 0)
        self.txt_obs = QLineEdit()
        g.addWidget(self.txt_obs, 3, 1, 1, 3)

        # Fila condiciones
        g.addWidget(QLabel("Moneda *"), 4, 0)
        self.cmb_moneda = QComboBox()
        g.addWidget(self.cmb_moneda, 4, 1)

        g.addWidget(QLabel("N° Letras (0=auto)"), 4, 2)
        self.spn_nletras = QSpinBox(); self.spn_nletras.setRange(0, 50)
        g.addWidget(self.spn_nletras, 4, 3)

        g.addWidget(QLabel("Fecha Emisión *"), 5, 0)
        self.dte_emi = QDateEdit(QDate.currentDate())
        self.dte_emi.setCalendarPopup(True); self.dte_emi.setDisplayFormat("dd/MM/yyyy")
        self.dte_emi.dateChanged.connect(self._recalc_fecha_lim)
        g.addWidget(self.dte_emi, 5, 1)

        g.addWidget(QLabel("Días Crédito *"), 5, 2)
        self.spn_dias = QSpinBox(); self.spn_dias.setRange(1, 3650)
        self.spn_dias.setValue(60)
        self.spn_dias.valueChanged.connect(self._recalc_fecha_lim)
        g.addWidget(self.spn_dias, 5, 3)

        g.addWidget(QLabel("Fecha Límite"), 6, 0)
        self.dte_lim = QDateEdit(QDate.currentDate())
        self.dte_lim.setCalendarPopup(True); self.dte_lim.setDisplayFormat("dd/MM/yyyy")
        g.addWidget(self.dte_lim, 6, 1)
        g.addWidget(QLabel("(editable — sobreescribe el cálculo)"), 6, 2, 1, 2)

        root.addWidget(grp_op)
        self._recalc_fecha_lim()

        # ── Tabla de facturas ─────────────────────────────────────────────────
        grp_fact = QGroupBox("Tabla de Facturas (hasta 50)")
        vf = QVBoxLayout(grp_fact)

        self.tbl_fact = QTableWidget(50, 4)
        self.tbl_fact.setHorizontalHeaderLabels(["N° Factura / Doc", "Importe", "Retención", "Neto"])
        self.tbl_fact.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        for c in (1, 2, 3):
            self.tbl_fact.horizontalHeader().setSectionResizeMode(c, QHeaderView.ResizeToContents)
        self.tbl_fact.setMaximumHeight(200)
        self.tbl_fact.itemChanged.connect(self._on_fact_changed)
        vf.addWidget(self.tbl_fact)

        hres = QHBoxLayout()
        hres.addWidget(QLabel("Bruto:"))
        self.lbl_bruto = QLabel("0.00"); self.lbl_bruto.setStyleSheet("font-weight:bold")
        hres.addWidget(self.lbl_bruto)
        hres.addSpacing(20)
        hres.addWidget(QLabel("Retención Total:"))
        self.lbl_ret = QLabel("0.00"); self.lbl_ret.setStyleSheet("font-weight:bold")
        hres.addWidget(self.lbl_ret)
        hres.addSpacing(20)
        hres.addWidget(QLabel("NETO TOTAL:"))
        self.lbl_neto = QLabel("0.00"); self.lbl_neto.setStyleSheet("font-weight:bold;font-size:13px;color:#1a5276")
        hres.addWidget(self.lbl_neto)
        hres.addStretch()
        vf.addLayout(hres)
        root.addWidget(grp_fact)

        # ── Distribución manual ───────────────────────────────────────────────
        grp_dist = QGroupBox("Distribución Manual por Período (opcional)")
        grp_dist.setCheckable(True); grp_dist.setChecked(False)
        vd = QVBoxLayout(grp_dist)
        vd.addWidget(QLabel("Ingrese un límite máximo de letras por mes para esta operación:"))
        self.tbl_dist = QTableWidget(11, 2)
        self.tbl_dist.setHorizontalHeaderLabels(["Mes (YYYY-MM)", "Importe Máx Esta Op."])
        self.tbl_dist.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tbl_dist.setMaximumHeight(180)
        vd.addWidget(self.tbl_dist)
        root.addWidget(grp_dist)
        self.grp_dist = grp_dist

        # ── Botones ───────────────────────────────────────────────────────────
        hbtn = QHBoxLayout()
        self.btn_gen   = QPushButton("⚡ Generar Cronograma")
        self.btn_gen.setFixedHeight(36)
        self.btn_gen.setStyleSheet("font-weight:bold;background:#1a5276;color:white")
        self.btn_gen.clicked.connect(self._generar)
        self.btn_limp  = QPushButton("🗑 Limpiar Formulario")
        self.btn_limp.setFixedHeight(36)
        self.btn_limp.clicked.connect(self._limpiar)
        hbtn.addWidget(self.btn_gen)
        hbtn.addWidget(self.btn_limp)
        hbtn.addStretch()
        root.addLayout(hbtn)

        # ── Cronograma resultado ──────────────────────────────────────────────
        grp_cron = QGroupBox("Cronograma Generado")
        vc = QVBoxLayout(grp_cron)

        hinfo = QHBoxLayout()
        hinfo.addWidget(QLabel("Total letras:"))
        self.lbl_nletras = QLabel("—"); hinfo.addWidget(self.lbl_nletras)
        hinfo.addSpacing(20)
        hinfo.addWidget(QLabel("Suma cronograma:"))
        self.lbl_suma = QLabel("—"); hinfo.addWidget(self.lbl_suma)
        hinfo.addSpacing(20)
        hinfo.addWidget(QLabel("Diferencia:"))
        self.lbl_diff = QLabel("—"); hinfo.addWidget(self.lbl_diff)
        hinfo.addStretch()
        self.btn_export = QPushButton("📊 Exportar Excel")
        self.btn_export.clicked.connect(self._exportar)
        self.btn_export.setEnabled(False)
        hinfo.addWidget(self.btn_export)
        self.btn_print = QPushButton("🖨 Imprimir / Vista Previa")
        self.btn_print.clicked.connect(self._imprimir)
        self.btn_print.setEnabled(False)
        hinfo.addWidget(self.btn_print)
        vc.addLayout(hinfo)

        self.tbl_cron = QTableWidget(0, 8)
        self.tbl_cron.setHorizontalHeaderLabels([
            "N°", "Código", "F. Emisión", "F. Vcto.", "Moneda", "Pres.Disp.", "Monto", "Estado"])
        self.tbl_cron.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        for c in (0, 2, 3, 4, 5, 6, 7):
            self.tbl_cron.horizontalHeader().setSectionResizeMode(c, QHeaderView.ResizeToContents)
        self.tbl_cron.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl_cron.setSelectionBehavior(QAbstractItemView.SelectRows)
        vc.addWidget(self.tbl_cron)
        root.addWidget(grp_cron)

    def _recalc_fecha_lim(self):
        qd = self.dte_emi.date()
        d  = date(qd.year(), qd.month(), qd.day())
        lim = d + timedelta(days=self.spn_dias.value())
        self.dte_lim.setDate(QDate(lim.year, lim.month, lim.day))

    def _on_prov_change(self, txt):
        base = self._get_data()[1]   # params_base
        for p in base.get("proveedores", []):
            if p["nombre"] == txt:
                self.txt_ruc.setText(p.get("ruc", ""))
                return
        self.txt_ruc.clear()

    def _on_fact_changed(self, item):
        self.tbl_fact.blockSignals(True)
        row = item.row()
        col = item.column()
        if col in (1, 2):
            try:
                imp_txt = self.tbl_fact.item(row, 1)
                ret_txt = self.tbl_fact.item(row, 2)
                imp = float((imp_txt.text() if imp_txt else "0").replace(",", ".") or 0)
                ret = float((ret_txt.text() if ret_txt else "0").replace(",", ".") or 0)
                neto = imp - ret
                it = QTableWidgetItem(f"{neto:,.2f}")
                it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                it.setFlags(it.flags() & ~Qt.ItemIsEditable)
                self.tbl_fact.setItem(row, 3, it)
            except Exception:
                pass
        self.tbl_fact.blockSignals(False)
        self._update_totales()

    def _update_totales(self):
        bruto = ret = 0.0
        for r in range(50):
            try:
                imp = float((self.tbl_fact.item(r, 1).text() if self.tbl_fact.item(r, 1) else "0").replace(",", ".") or 0)
                rr  = float((self.tbl_fact.item(r, 2).text() if self.tbl_fact.item(r, 2) else "0").replace(",", ".") or 0)
                bruto += imp; ret += rr
            except Exception:
                pass
        neto = bruto - ret
        self.lbl_bruto.setText(f"{bruto:,.2f}")
        self.lbl_ret.setText(f"{ret:,.2f}")
        self.lbl_neto.setText(f"{neto:,.2f}")

    def _get_facturas(self):
        facts = []
        for r in range(50):
            try:
                doc = (self.tbl_fact.item(r, 0).text() if self.tbl_fact.item(r, 0) else "").strip()
                imp = float((self.tbl_fact.item(r, 1).text() if self.tbl_fact.item(r, 1) else "").replace(",", ".") or 0)
                ret = float((self.tbl_fact.item(r, 2).text() if self.tbl_fact.item(r, 2) else "").replace(",", ".") or 0)
                if imp > 0 or doc:
                    facts.append({"doc": doc, "importe": imp, "retencion": ret})
            except Exception:
                pass
        return facts

    def _get_dist(self):
        dist = []
        if not self.grp_dist.isChecked():
            return dist
        for r in range(11):
            try:
                mes = (self.tbl_dist.item(r, 0).text() if self.tbl_dist.item(r, 0) else "").strip()
                lim = float((self.tbl_dist.item(r, 1).text() if self.tbl_dist.item(r, 1) else "").replace(",", ".") or 0)
                if mes and lim > 0:
                    dist.append({"mes": mes, "limite": lim})
            except Exception:
                pass
        return dist

    def _generar(self):
        empresa  = self.cmb_empresa.currentText().strip()
        temporada= self.cmb_temp.currentText().strip()
        proveedor= self.cmb_prov.currentText().strip()
        ruc      = self.txt_ruc.text().strip()
        concepto = self.txt_concepto.text().strip()
        observac = self.txt_obs.text().strip()
        moneda   = self.cmb_moneda.currentText().split()[0] if self.cmb_moneda.currentText() else "PEN"
        n_letras = self.spn_nletras.value()

        qd_emi = self.dte_emi.date()
        qd_lim = self.dte_lim.date()
        fecha_emi = date(qd_emi.year(), qd_emi.month(), qd_emi.day())
        fecha_lim = date(qd_lim.year(), qd_lim.month(), qd_lim.day())
        dias_cred = self.spn_dias.value()

        errors = []
        if not empresa:   errors.append("Empresa Emisora requerida.")
        if not proveedor: errors.append("Proveedor requerido.")
        if not concepto:  errors.append("Concepto requerido.")
        if fecha_lim <= fecha_emi: errors.append("Fecha Límite debe ser posterior a Fecha Emisión.")
        if errors:
            QMessageBox.warning(self, "Campos requeridos", "\n".join(errors))
            return

        facturas = self._get_facturas()
        if not facturas:
            QMessageBox.warning(self, "Sin facturas", "Ingrese al menos una factura con importe.")
            return

        dist_manual = self._get_dist()
        params, params_base, historico, externos = self._get_data()

        try:
            letras, avisos = generar_cronograma(
                empresa, temporada, proveedor, ruc, concepto, observac, moneda,
                n_letras, fecha_emi, dias_cred, fecha_lim,
                facturas, dist_manual,
                params, params_base, historico, externos
            )
        except ValueError as e:
            QMessageBox.critical(self, "Error", str(e))
            return

        self._letras = letras
        self._mostrar_cronograma(letras)

        if avisos:
            QMessageBox.information(self, "Avisos", "\n".join(avisos))

        # Confirmar grabación en histórico
        resp = QMessageBox.question(
            self, "Guardar en Histórico",
            f"Se generaron {len(letras)} letra(s).\n¿Guardar en el Histórico?",
            QMessageBox.Yes | QMessageBox.No
        )
        if resp == QMessageBox.Yes:
            self.cronograma_generado.emit(letras)
            QMessageBox.information(self, "Listo", "Letras guardadas en el Histórico.")

    def _mostrar_cronograma(self, letras):
        self.tbl_cron.setRowCount(len(letras))
        for i, lt in enumerate(letras):
            moneda = lt["moneda"]
            self.tbl_cron.setItem(i, 0, num_item(lt["n"], "{:.0f}", Qt.AlignCenter | Qt.AlignVCenter))
            self.tbl_cron.setItem(i, 1, make_item(lt["codigo"]))
            self.tbl_cron.setItem(i, 2, make_item(lt["fecha_emi"], Qt.AlignCenter | Qt.AlignVCenter))
            self.tbl_cron.setItem(i, 3, make_item(lt["fecha_vcto"], Qt.AlignCenter | Qt.AlignVCenter))
            self.tbl_cron.setItem(i, 4, make_item(moneda, Qt.AlignCenter | Qt.AlignVCenter))
            self.tbl_cron.setItem(i, 5, num_item(lt["espacio"]))
            self.tbl_cron.setItem(i, 6, num_item(lt["monto"]))
            estado = lt["estado"]
            it_est = make_item(estado, Qt.AlignCenter | Qt.AlignVCenter)
            color = {"Parcial": "#f39c12", "Ajuste": "#e74c3c"}.get(estado, "")
            if color:
                it_est.setForeground(QBrush(QColor(color)))
            self.tbl_cron.setItem(i, 7, it_est)

        suma = sum(lt["monto"] for lt in letras)
        moneda0 = letras[0]["moneda"] if letras else ""
        neto_total = sum(float((self.tbl_fact.item(r,1).text() if self.tbl_fact.item(r,1) else "0").replace(",",".") or 0) -
                         float((self.tbl_fact.item(r,2).text() if self.tbl_fact.item(r,2) else "0").replace(",",".") or 0)
                         for r in range(50))
        diff = neto_total - suma

        self.lbl_nletras.setText(str(len(letras)))
        self.lbl_suma.setText(f"{moneda0} {suma:,.2f}")
        self.lbl_diff.setText(f"{diff:,.2f}")
        color_diff = "red" if abs(diff) > 0.1 else "#27ae60"
        self.lbl_diff.setStyleSheet(f"color:{color_diff};font-weight:bold")
        self.btn_export.setEnabled(True)
        self.btn_print.setEnabled(True)

    def _limpiar(self):
        resp = QMessageBox.question(self, "Limpiar", "¿Limpiar el formulario? El Histórico se conserva.",
                                    QMessageBox.Yes | QMessageBox.No)
        if resp != QMessageBox.Yes:
            return
        self.txt_concepto.clear(); self.txt_obs.clear()
        self.spn_nletras.setValue(0); self.spn_dias.setValue(60)
        self.tbl_fact.clearContents()
        self.tbl_dist.clearContents()
        self.tbl_cron.setRowCount(0)
        self._letras = []
        self.btn_export.setEnabled(False)
        self.btn_print.setEnabled(False)
        self.lbl_bruto.setText("0.00"); self.lbl_ret.setText("0.00")
        self.lbl_neto.setText("0.00"); self.lbl_nletras.setText("—")
        self.lbl_suma.setText("—"); self.lbl_diff.setText("—")

    def _exportar(self):
        if not self._letras:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from PySide6.QtWidgets import QFileDialog
            path, _ = QFileDialog.getSaveFileName(
                self, "Guardar Excel", str(BASE / "cronograma.xlsx"),
                "Excel (*.xlsx)")
            if not path:
                return
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Cronograma"

            # Info operación
            lt0 = self._letras[0]
            ws.append(["CRONOGRAMA DE LETRAS"])
            ws.append(["Empresa:", lt0["empresa"], "", "Proveedor:", lt0["proveedor"]])
            ws.append(["Concepto:", lt0["concepto"], "", "Temporada:", lt0["temporada"]])
            ws.append([])
            headers = ["N°", "Código", "F. Emisión", "F. Vcto.", "Moneda", "Monto", "Estado"]
            ws.append(headers)
            for lt in self._letras:
                ws.append([lt["n"], lt["codigo"], lt["fecha_emi"], lt["fecha_vcto"],
                           lt["moneda"], lt["monto"], lt["estado"]])
            ws.append([])
            ws.append(["", "TOTAL", "", "", "", sum(l["monto"] for l in self._letras)])

            # Facturas
            ws2 = wb.create_sheet("Facturas")
            ws2.append(["N° Factura", "Importe", "Retención", "Neto"])
            for r in range(50):
                try:
                    doc = (self.tbl_fact.item(r,0).text() if self.tbl_fact.item(r,0) else "").strip()
                    imp = float((self.tbl_fact.item(r,1).text() if self.tbl_fact.item(r,1) else "").replace(",",".") or 0)
                    ret = float((self.tbl_fact.item(r,2).text() if self.tbl_fact.item(r,2) else "").replace(",",".") or 0)
                    if imp > 0 or doc:
                        ws2.append([doc, imp, ret, imp - ret])
                except Exception:
                    pass
            wb.save(path)
            QMessageBox.information(self, "Exportado", f"Archivo guardado:\n{path}")
        except ImportError:
            QMessageBox.critical(self, "Error", "Instala openpyxl:\n\npip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    # ── Imprimible ────────────────────────────────────────────────────────────
    def _build_html_imprimible(self):
        """Genera el HTML del documento imprimible, fiel al diseño del Excel."""
        if not self._letras:
            return ""
        lt0 = self._letras[0]
        empresa   = lt0.get("empresa",   "")
        proveedor = lt0.get("proveedor", "")
        ruc       = lt0.get("ruc",       "")
        concepto  = lt0.get("concepto",  "")
        observac  = lt0.get("observac",  "")
        temporada = lt0.get("temporada", "")
        moneda    = lt0.get("moneda",    "PEN")
        fecha_emi = lt0.get("fecha_emi", "")
        dias_cred = self.spn_dias.value()
        qd_lim    = self.dte_lim.date()
        fecha_lim = f"{qd_lim.day():02d}/{qd_lim.month():02d}/{qd_lim.year()}"

        # Facturas
        facturas = self._get_facturas()
        bruto = sum(f["importe"] for f in facturas)
        ret_t = sum(f["retencion"] for f in facturas)
        neto  = bruto - ret_t

        # Letras
        suma_cron = sum(lt["monto"] for lt in self._letras)
        moneda_sym = {"PEN": "S/", "USD": "$", "EUR": "€"}.get(moneda, moneda)

        def fmt_n(v):
            try: return f"{float(v):,.2f}"
            except: return str(v)

        def fmt_d(iso):
            try:
                y, m, d = iso.split("-")
                return f"{d}/{m}/{y}"
            except:
                return iso

        # ─── estilos ────────────────────────────────────────────────────────
        css = """
        body { font-family: Calibri, Arial, sans-serif; font-size: 9pt;
               margin: 0; padding: 0; color: #000; }
        .page  { width: 740px; margin: 0 auto; }
        h1 { font-size: 13pt; font-weight: bold; text-align: center;
             color: #1a3a5c; margin: 0 0 2px 0; }
        h2 { font-size: 9pt; font-weight: bold; text-align: center;
             color: #1a3a5c; margin: 0 0 8px 0; letter-spacing: 1px; }
        .seccion { background:#1a3a5c; color:white; font-weight:bold;
                   padding:3px 6px; font-size:8.5pt; margin-top:10px; }
        table { border-collapse: collapse; width: 100%; font-size: 8.5pt; }
        th { background:#1a3a5c; color:white; font-weight:bold;
             padding:4px 6px; text-align:center; border:1px solid #aaa; }
        td { padding:3px 6px; border:1px solid #ccc; }
        td.num { text-align:right; }
        td.cen { text-align:center; }
        .tot  { background:#dde8f0; font-weight:bold; }
        .lbl  { color:#555; font-size:8pt; }
        .val  { font-weight:bold; }
        .footer { font-size:7.5pt; color:#555; text-align:center;
                  margin-top:14px; border-top:1px solid #ccc; padding-top:4px; }
        .info-grid { width:100%; border-collapse:collapse; margin-bottom:6px; }
        .info-grid td { border:none; padding:2px 4px; font-size:8.5pt; }
        .firma-row { margin-top:40px; width:100%; }
        .firma-box { display:inline-block; text-align:center; width:200px;
                     border-top:1px solid #000; padding-top:4px; font-size:8pt; }
        """

        # ─── cabecera operación ──────────────────────────────────────────────
        try:
            fe_fmt = fmt_d(fecha_emi)
        except:
            fe_fmt = fecha_emi

        html_cab = f"""
        <table class="info-grid">
          <tr>
            <td class="lbl">Proveedor:</td>
            <td class="val" colspan="3">{proveedor}{"  —  " + ruc if ruc else ""}</td>
          </tr>
          <tr>
            <td class="lbl">Fecha Emisión:</td>
            <td class="val">{fe_fmt}</td>
            <td class="lbl">Días Crédito:</td>
            <td class="val">{dias_cred} días</td>
          </tr>
          <tr>
            <td class="lbl">Fecha Límite:</td>
            <td class="val">{fecha_lim}</td>
            <td class="lbl">Temporada:</td>
            <td class="val">{temporada}</td>
          </tr>
          <tr>
            <td class="lbl">Concepto:</td>
            <td class="val" colspan="3">{concepto}</td>
          </tr>"""
        if observac:
            html_cab += f"""
          <tr>
            <td class="lbl">Observaciones:</td>
            <td class="val" colspan="3">{observac}</td>
          </tr>"""
        html_cab += "</table>"

        # ─── tabla facturas ──────────────────────────────────────────────────
        rows_fact = ""
        for i, f in enumerate(facturas):
            pct = f["retencion"] / f["importe"] * 100 if f["importe"] else 0
            rows_fact += f"""<tr>
              <td class="cen">{i+1}</td>
              <td>{f['doc']}</td>
              <td class="num">{fmt_n(f['importe'])}</td>
              <td class="cen">{pct:.0f}%</td>
              <td class="num">{fmt_n(f['retencion'])}</td>
              <td class="num">{fmt_n(f['importe']-f['retencion'])}</td>
            </tr>"""
        html_fact = f"""
        <div class="seccion">FACTURAS REFERENCIADAS</div>
        <table>
          <tr>
            <th style="width:30px">N</th>
            <th>N° FACTURA / DOC</th>
            <th style="width:90px">IMPORTE</th>
            <th style="width:40px">RET.</th>
            <th style="width:100px">MTO. RETENCIÓN</th>
            <th style="width:90px">NETO</th>
          </tr>
          {rows_fact}
          <tr class="tot">
            <td colspan="2" style="text-align:right">TOTALES</td>
            <td class="num">{fmt_n(bruto)}</td>
            <td></td>
            <td class="num">{fmt_n(ret_t)}</td>
            <td class="num">{fmt_n(neto)}</td>
          </tr>
        </table>"""

        # ─── tabla cronograma ────────────────────────────────────────────────
        rows_cron = ""
        for lt in self._letras:
            f_emi = fmt_d(lt["fecha_emi"])
            f_vct = fmt_d(lt["fecha_vcto"])
            try:
                dias = (date.fromisoformat(lt["fecha_vcto"]) - date.fromisoformat(lt["fecha_emi"])).days
            except:
                dias = ""
            rows_cron += f"""<tr>
              <td class="cen">{lt['n']}</td>
              <td class="cen">{lt['codigo']}</td>
              <td class="cen">{f_emi}</td>
              <td class="cen">{f_vct}</td>
              <td class="cen">{dias}</td>
              <td class="cen">{lt['moneda']}</td>
              <td class="num">{fmt_n(lt['monto'])}</td>
            </tr>"""
        html_cron = f"""
        <div class="seccion">CRONOGRAMA DE LETRAS</div>
        <table style="margin-top:4px">
          <tr style="background:#dde8f0; font-size:8pt;">
            <td colspan="2"><span class="lbl">N° de Letras:</span>
              <b>{len(self._letras)}</b></td>
            <td colspan="3"><span class="lbl">Neto Total:</span>
              <b>{moneda_sym} {fmt_n(neto)}</b></td>
            <td colspan="2"><span class="lbl">Suma Cronograma:</span>
              <b>{moneda_sym} {fmt_n(suma_cron)}</b></td>
          </tr>
          <tr>
            <th style="width:30px">N°</th>
            <th>CÓDIGO LETRA</th>
            <th style="width:80px">FECHA EMI.</th>
            <th style="width:80px">FECHA VCTO.</th>
            <th style="width:35px">DÍAS</th>
            <th style="width:55px">MONEDA</th>
            <th style="width:90px">MONTO LETRA</th>
          </tr>
          {rows_cron}
          <tr class="tot">
            <td colspan="6" style="text-align:right">TOTAL CRONOGRAMA</td>
            <td class="num">{moneda_sym} {fmt_n(suma_cron)}</td>
          </tr>
        </table>"""

        # ─── firmas ──────────────────────────────────────────────────────────
        html_firmas = """
        <table style="margin-top:50px; border:none; width:100%">
          <tr>
            <td style="width:33%; text-align:center; border:none; padding-top:30px">
              <div style="border-top:1px solid #000; display:inline-block;
                          width:180px; padding-top:4px; font-size:8pt">
                Elaborado por
              </div>
            </td>
            <td style="width:33%; text-align:center; border:none; padding-top:30px">
              <div style="border-top:1px solid #000; display:inline-block;
                          width:180px; padding-top:4px; font-size:8pt">
                Revisado por
              </div>
            </td>
            <td style="width:33%; text-align:center; border:none; padding-top:30px">
              <div style="border-top:1px solid #000; display:inline-block;
                          width:180px; padding-top:4px; font-size:8pt">
                Autorizado por
              </div>
            </td>
          </tr>
        </table>"""

        # último vencimiento
        try:
            ult_vcto = max(date.fromisoformat(lt["fecha_vcto"]) for lt in self._letras)
            ult_vcto_str = ult_vcto.strftime("%d/%m/%Y")
        except:
            ult_vcto_str = ""

        footer = f"""
        <div class="footer">
          Emitido: {fe_fmt} &nbsp;|&nbsp;
          Vencimiento último: {ult_vcto_str} &nbsp;|&nbsp;
          Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}
        </div>"""

        return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>{css}</style></head>
<body>
<div class="page">
  <h1>{empresa}</h1>
  <h2>LIQUIDACIÓN DE FACTURAS CANJEADAS A LETRAS</h2>
  {html_cab}
  {html_fact}
  {html_cron}
  {html_firmas}
  {footer}
</div>
</body></html>"""

    def _imprimir(self):
        if not self._letras:
            return
        if not _PRINT_OK:
            QMessageBox.critical(self, "Error",
                "Módulo de impresión no disponible.\n"
                "Instale: pip install PySide6-Addons")
            return
        html = self._build_html_imprimible()
        printer = QPrinter(QPrinter.HighResolution)
        printer.setPageSize(QPageSize(QPageSize.A4))
        printer.setPageOrientation(QPrinter.Portrait)

        doc = QTextDocument()
        doc.setHtml(html)
        doc.setPageSize(QSizeF(printer.pageRect(QPrinter.DevicePixel).size()))

        preview = QPrintPreviewDialog(printer, self)
        preview.setWindowTitle("Vista Previa — Liquidación de Letras")
        preview.resize(900, 700)

        def render(p):
            doc.print_(p)

        preview.paintRequested.connect(render)
        preview.exec()

    def refresh_combos(self, base):
        """Actualiza combos desde base de datos."""
        cur_emp  = self.cmb_empresa.currentText()
        cur_prov = self.cmb_prov.currentText()
        cur_temp = self.cmb_temp.currentText()
        cur_mon  = self.cmb_moneda.currentText()

        self.cmb_empresa.clear()
        self.cmb_empresa.addItems(base.get("empresas", []))
        if cur_emp: self.cmb_empresa.setCurrentText(cur_emp)

        self.cmb_prov.clear()
        self.cmb_prov.addItems([p["nombre"] for p in base.get("proveedores", [])])
        if cur_prov: self.cmb_prov.setCurrentText(cur_prov)

        self.cmb_temp.clear()
        self.cmb_temp.addItems(base.get("temporadas", []))
        if cur_temp: self.cmb_temp.setCurrentText(cur_temp)

        self.cmb_moneda.clear()
        for m in base.get("monedas", []):
            self.cmb_moneda.addItem(f"{m['codigo']} — {m['nombre']}", m["codigo"])
        if cur_mon: self.cmb_moneda.setCurrentText(cur_mon)


# ── Tab: Parámetros ────────────────────────────────────────────────────────────
class TabParametros(QWidget):
    params_changed = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self._params = load_params()
        self._build()
        self._load_ui()

    def _build(self):
        root = QVBoxLayout(self)
        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        inner = QWidget(); root.addWidget(scroll); scroll.setWidget(inner)
        v = QVBoxLayout(inner); v.setSpacing(8)

        # Config general
        grp_gen = QGroupBox("Configuración General")
        fg = QFormLayout(grp_gen)
        self.spn_presup_def = QDoubleSpinBox(); self.spn_presup_def.setRange(0, 1e9); self.spn_presup_def.setDecimals(2); self.spn_presup_def.setSuffix(" PEN")
        self.spn_sep_min    = QSpinBox();       self.spn_sep_min.setRange(0, 365)
        self.spn_monto_min  = QDoubleSpinBox(); self.spn_monto_min.setRange(0, 1e9); self.spn_monto_min.setDecimals(2)
        self.spn_monto_max  = QDoubleSpinBox(); self.spn_monto_max.setRange(0, 1e9); self.spn_monto_max.setDecimals(2)
        self.spn_lim_max    = QDoubleSpinBox(); self.spn_lim_max.setRange(0, 1e9); self.spn_lim_max.setDecimals(2)
        self.spn_parcial_min= QDoubleSpinBox(); self.spn_parcial_min.setRange(0, 1e9); self.spn_parcial_min.setDecimals(2)
        self.spn_parcial_max= QDoubleSpinBox(); self.spn_parcial_max.setRange(0, 1e9); self.spn_parcial_max.setDecimals(2)
        fg.addRow("Presupuesto mensual por defecto:", self.spn_presup_def)
        fg.addRow("Separación mínima entre letras (días):", self.spn_sep_min)
        fg.addRow("Monto mínimo por letra (PEN):", self.spn_monto_min)
        fg.addRow("Monto máximo por letra (PEN):", self.spn_monto_max)
        fg.addRow("Límite diario máximo (PEN, 0=sin límite):", self.spn_lim_max)
        fg.addRow("Monto mínimo letra parcial (PEN):", self.spn_parcial_min)
        fg.addRow("Monto máximo letra parcial (PEN):", self.spn_parcial_max)
        v.addWidget(grp_gen)

        # Días bloqueados
        grp_dias = QGroupBox("Días No Hábiles")
        gd = QHBoxLayout(grp_dias)
        self.chk_dias = {}
        for dia in ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"]:
            chk = QCheckBox(dia)
            self.chk_dias[dia] = chk
            gd.addWidget(chk)
        gd.addStretch()
        v.addWidget(grp_dias)

        # Períodos bloqueados
        grp_per = QGroupBox("Períodos / Fechas Específicas Bloqueadas")
        vp = QVBoxLayout(grp_per)
        self.tbl_per = QTableWidget(20, 3)
        self.tbl_per.setHorizontalHeaderLabels(["Descripción", "Fecha Inicio (YYYY-MM-DD)", "Fecha Fin (YYYY-MM-DD)"])
        self.tbl_per.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.tbl_per.setMaximumHeight(200)
        vp.addWidget(self.tbl_per)
        v.addWidget(grp_per)

        # Presupuesto mensual
        grp_pres = QGroupBox("Presupuesto Disponible por Período")
        vpres = QVBoxLayout(grp_pres)
        vpres.addWidget(QLabel("Formato mes: YYYY-MM (ej: 2026-04). Override sobreescribe el límite diario calculado."))
        self.tbl_pres = QTableWidget(36, 3)
        self.tbl_pres.setHorizontalHeaderLabels(["Mes (YYYY-MM)", "Presupuesto Disponible", "Override Límite Diario"])
        self.tbl_pres.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        vpres.addWidget(self.tbl_pres)
        v.addWidget(grp_pres)

        # Monedas
        grp_mon = QGroupBox("Tipos de Cambio (referencia)")
        vm = QVBoxLayout(grp_mon)
        self.tbl_mon = QTableWidget(10, 3)
        self.tbl_mon.setHorizontalHeaderLabels(["Código", "Nombre", "Tipo de Cambio"])
        self.tbl_mon.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.tbl_mon.setMaximumHeight(200)
        vm.addWidget(self.tbl_mon)
        v.addWidget(grp_mon)

        btn = QPushButton("💾 Guardar Parámetros")
        btn.setFixedHeight(36); btn.setStyleSheet("font-weight:bold")
        btn.clicked.connect(self._guardar)
        v.addWidget(btn)

    def _load_ui(self):
        p = self._params
        self.spn_presup_def.setValue(float(p.get("presupuesto_default", 0) or 0))
        self.spn_sep_min.setValue(int(p.get("separacion_minima", 0) or 0))
        self.spn_monto_min.setValue(float(p.get("monto_minimo", 0) or 0))
        self.spn_monto_max.setValue(float(p.get("monto_maximo", 0) or 0))
        self.spn_lim_max.setValue(float(p.get("limite_diario_max", 0) or 0))
        self.spn_parcial_min.setValue(float(p.get("parcial_minimo", 0) or 0))
        self.spn_parcial_max.setValue(float(p.get("parcial_maximo", 0) or 0))

        for dia, chk in self.chk_dias.items():
            chk.setChecked(p.get("dias_bloqueados", {}).get(dia, False))

        for i, per in enumerate(p.get("periodos_bloqueados", [])[:20]):
            self.tbl_per.setItem(i, 0, QTableWidgetItem(per.get("descripcion", "")))
            self.tbl_per.setItem(i, 1, QTableWidgetItem(per.get("inicio", "")))
            self.tbl_per.setItem(i, 2, QTableWidgetItem(per.get("fin", "")))

        for i, row in enumerate(p.get("presupuesto_mensual", [])[:36]):
            self.tbl_pres.setItem(i, 0, QTableWidgetItem(row.get("mes", "")))
            self.tbl_pres.setItem(i, 1, QTableWidgetItem(str(row.get("presupuesto", ""))))
            self.tbl_pres.setItem(i, 2, QTableWidgetItem(str(row.get("override", ""))))

        base = load_base()
        for i, m in enumerate(base.get("monedas", [])[:10]):
            self.tbl_mon.setItem(i, 0, QTableWidgetItem(m.get("codigo", "")))
            self.tbl_mon.setItem(i, 1, QTableWidgetItem(m.get("nombre", "")))
            self.tbl_mon.setItem(i, 2, QTableWidgetItem(str(m.get("tc", 1.0))))

    def _guardar(self):
        p = self._params
        p["presupuesto_default"]  = self.spn_presup_def.value()
        p["separacion_minima"]    = self.spn_sep_min.value()
        p["monto_minimo"]         = self.spn_monto_min.value()
        p["monto_maximo"]         = self.spn_monto_max.value()
        p["limite_diario_max"]    = self.spn_lim_max.value()
        p["parcial_minimo"]       = self.spn_parcial_min.value()
        p["parcial_maximo"]       = self.spn_parcial_max.value()
        p["dias_bloqueados"]      = {dia: chk.isChecked() for dia, chk in self.chk_dias.items()}

        periodos = []
        for i in range(20):
            desc = (self.tbl_per.item(i,0).text() if self.tbl_per.item(i,0) else "").strip()
            ini  = (self.tbl_per.item(i,1).text() if self.tbl_per.item(i,1) else "").strip()
            fin  = (self.tbl_per.item(i,2).text() if self.tbl_per.item(i,2) else "").strip()
            if ini and fin:
                periodos.append({"descripcion": desc, "inicio": ini, "fin": fin})
        p["periodos_bloqueados"] = periodos

        pres_rows = []
        for i in range(36):
            mes = (self.tbl_pres.item(i,0).text() if self.tbl_pres.item(i,0) else "").strip()
            if not mes: continue
            try:
                presup  = float((self.tbl_pres.item(i,1).text() if self.tbl_pres.item(i,1) else "").replace(",",".") or 0)
                override= float((self.tbl_pres.item(i,2).text() if self.tbl_pres.item(i,2) else "").replace(",",".") or 0)
                pres_rows.append({"mes": mes, "presupuesto": presup, "override": override})
            except Exception:
                pass
        p["presupuesto_mensual"] = pres_rows
        save_params(p)

        # Monedas → base
        base = load_base()
        monedas = []
        for i in range(10):
            cod  = (self.tbl_mon.item(i,0).text() if self.tbl_mon.item(i,0) else "").strip()
            if not cod: continue
            nom  = (self.tbl_mon.item(i,1).text() if self.tbl_mon.item(i,1) else "").strip()
            try:
                tc = float((self.tbl_mon.item(i,2).text() if self.tbl_mon.item(i,2) else "1").replace(",",".") or 1)
            except Exception:
                tc = 1.0
            monedas.append({"codigo": cod, "nombre": nom, "tc": tc})
        base["monedas"] = monedas
        save_base(base)

        QMessageBox.information(self, "Guardado", "Parámetros guardados correctamente.")
        self.params_changed.emit()

    def get_params(self):
        return self._params


# ── Tab: Histórico ─────────────────────────────────────────────────────────────
class TabHistorico(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._historico = []
        self._build()

    def _build(self):
        v = QVBoxLayout(self)

        h = QHBoxLayout()
        h.addWidget(QLabel("Buscar:"))
        self.txt_buscar = QLineEdit(); self.txt_buscar.setPlaceholderText("Proveedor, código, concepto...")
        self.txt_buscar.textChanged.connect(self._filtrar)
        h.addWidget(self.txt_buscar)
        btn_del = QPushButton("Eliminar operación seleccionada")
        btn_del.clicked.connect(self._eliminar_op)
        h.addWidget(btn_del)
        v.addLayout(h)

        self.tbl = QTableWidget(0, 9)
        self.tbl.setHorizontalHeaderLabels([
            "ID Op.", "Fecha Reg.", "Empresa", "Proveedor", "Temporada",
            "N° Letra", "Moneda", "F. Vcto.", "Monto"])
        self.tbl.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        for c in (0,1,2,4,5,6,7,8):
            self.tbl.horizontalHeader().setSectionResizeMode(c, QHeaderView.ResizeToContents)
        self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
        v.addWidget(self.tbl)

        self.lbl_total = QLabel("Total: 0 letras")
        v.addWidget(self.lbl_total)

    def refresh(self, historico: list):
        self._historico = historico
        self._filtrar(self.txt_buscar.text())

    def _filtrar(self, texto: str):
        txt = texto.lower()
        rows = [h for h in self._historico
                if not txt or any(txt in str(v).lower() for v in h.values())]
        self.tbl.setRowCount(len(rows))
        for i, h in enumerate(rows):
            self.tbl.setItem(i, 0, make_item(h.get("id_operacion", "")))
            self.tbl.setItem(i, 1, make_item(h.get("fecha_registro", "")[:16]))
            self.tbl.setItem(i, 2, make_item(h.get("empresa", "")))
            self.tbl.setItem(i, 3, make_item(h.get("proveedor", "")))
            self.tbl.setItem(i, 4, make_item(h.get("temporada", "")))
            self.tbl.setItem(i, 5, make_item(str(h.get("n_letra", "")), Qt.AlignCenter | Qt.AlignVCenter))
            self.tbl.setItem(i, 6, make_item(h.get("moneda", ""), Qt.AlignCenter | Qt.AlignVCenter))
            self.tbl.setItem(i, 7, make_item(h.get("fecha_vcto", ""), Qt.AlignCenter | Qt.AlignVCenter))
            self.tbl.setItem(i, 8, num_item(h.get("monto", 0)))
            self.tbl.item(i, 0).setData(Qt.UserRole, h.get("id_operacion", ""))
        self.lbl_total.setText(f"Total: {len(rows)} letras")

    def _eliminar_op(self):
        row = self.tbl.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Sin selección", "Seleccione una fila.")
            return
        id_op = self.tbl.item(row, 0).data(Qt.UserRole)
        resp = QMessageBox.question(
            self, "Confirmar",
            f"¿Eliminar TODAS las letras de la operación {id_op}?",
            QMessageBox.Yes | QMessageBox.No)
        if resp != QMessageBox.Yes:
            return
        self._historico = [h for h in self._historico if h.get("id_operacion") != id_op]
        save_hist(self._historico)
        self._filtrar(self.txt_buscar.text())
        QMessageBox.information(self, "Eliminado", f"Operación {id_op} eliminada.")


# ── DYSPOW: Diálogo de Configuración ─────────────────────────────────────────
class DyspowConfigDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Conexión DYSPOW")
        self.setFixedWidth(380)
        cfg = load_dyspow_cfg()

        form = QFormLayout(self)
        self.txt_server   = QLineEdit(cfg.get("server", ""))
        self.txt_database = QLineEdit(cfg.get("database", ""))
        self.txt_user     = QLineEdit(cfg.get("user", ""))
        self.txt_password = QLineEdit(cfg.get("password", ""))
        self.txt_password.setEchoMode(QLineEdit.Password)

        form.addRow("Servidor:", self.txt_server)
        form.addRow("Base de Datos:", self.txt_database)
        form.addRow("Usuario:", self.txt_user)
        form.addRow("Contraseña:", self.txt_password)

        hb = QHBoxLayout()
        btn_ok  = QPushButton("Conectar")
        btn_can = QPushButton("Cancelar")
        btn_ok.clicked.connect(self.accept)
        btn_can.clicked.connect(self.reject)
        hb.addStretch(); hb.addWidget(btn_ok); hb.addWidget(btn_can)
        form.addRow(hb)

    def get_config(self):
        return {
            "server":   self.txt_server.text().strip(),
            "database": self.txt_database.text().strip(),
            "user":     self.txt_user.text().strip(),
            "password": self.txt_password.text(),
        }


def _dyspow_connect(cfg: dict):
    """Abre una conexión pyodbc a DYSPOW. Lanza excepción si falla."""
    if not _PYODBC_OK:
        raise RuntimeError("pyodbc no está instalado. Ejecute: pip install pyodbc")
    conn_str = (
        "DRIVER={ODBC Driver 17 for SQL Server};"
        f"SERVER={cfg['server']};"
        f"DATABASE={cfg['database']};"
        f"UID={cfg['user']};"
        f"PWD={cfg['password']};"
        "TrustServerCertificate=yes;"
    )
    return pyodbc.connect(conn_str, timeout=10)


# ── DYSPOW: Workers de importación ────────────────────────────────────────────
class WorkerDyspowProveedores(QThread):
    """Importa proveedores (nombre+RUC) y empresas desde DYSPOW."""
    done  = Signal(list, list)   # (proveedores, empresas)
    error = Signal(str)

    def __init__(self, cfg, parent=None):
        super().__init__(parent)
        self._cfg = cfg

    def run(self):
        try:
            conn = _dyspow_connect(self._cfg)
            cur  = conn.cursor()

            # Proveedores: razón social + RUC, activos, no vacíos
            cur.execute("""
                SELECT ISNULL(C_VAR_RAZON_SOCIAL, ''), ISNULL(C_VAR_NUMERO_DOCUMENTO, '')
                FROM T_060_PROVEEDOR
                WHERE C_BIT_ESTADO = 1
                  AND ISNULL(C_VAR_RAZON_SOCIAL, '') <> ''
                ORDER BY C_VAR_RAZON_SOCIAL
            """)
            provs = [{"nombre": r[0].strip(), "ruc": r[1].strip()} for r in cur.fetchall()]

            # Empresas: razón social
            cur.execute("""
                SELECT ISNULL(C_VAR_RAZON_SOCIAL, ''), ISNULL(C_VAR_NUMERO_DOCUMENTO, '')
                FROM T_019_EMPRESA
                WHERE ISNULL(C_VAR_RAZON_SOCIAL, '') <> ''
                ORDER BY C_INT_CODIGO_EMPRESA
            """)
            emps = [r[0].strip() for r in cur.fetchall()]

            conn.close()
            self.done.emit(provs, emps)
        except Exception as ex:
            self.error.emit(str(ex))


class WorkerDyspowLetras(QThread):
    """Importa letras de pago de DYSPOW → formato datos externos."""
    done  = Signal(list)   # lista de registros externos
    error = Signal(str)

    def __init__(self, cfg, params_base, parent=None):
        super().__init__(parent)
        self._cfg        = cfg
        self._params_base = params_base

    def run(self):
        try:
            conn = _dyspow_connect(self._cfg)
            cur  = conn.cursor()

            cur.execute("""
                SELECT
                    lp.C_VAR_NUMERO_LETRA_PAGO,
                    lp.C_DAT_FECHA_MOVIMIENTO,
                    lp.C_DEC_TOTAL,
                    ISNULL(mn.C_VAR_CODIGO_SUNAT, 'PEN') AS moneda_cod,
                    ISNULL(pr.C_VAR_RAZON_SOCIAL, '') AS proveedor,
                    ISNULL(emp.C_VAR_RAZON_SOCIAL, '') AS empresa,
                    ISNULL(lp.C_VAR_OBSERVACION, '') AS observacion
                FROM T_267_LETRA_PAGO lp
                LEFT JOIN T_066_MONEDA mn
                    ON mn.C_INT_CODIGO_EMPRESA = lp.C_INT_CODIGO_EMPRESA
                   AND mn.C_INT_CODIGO_MONEDA  = lp.C_INT_CODIGO_MONEDA
                LEFT JOIN T_060_PROVEEDOR pr
                    ON pr.C_INT_CODIGO_EMPRESA  = lp.C_INT_CODIGO_EMPRESA
                   AND pr.C_INT_CODIGO_PROVEEDOR = lp.C_INT_CODIGO_PROVEEDOR
                LEFT JOIN T_019_EMPRESA emp
                    ON emp.C_INT_CODIGO_EMPRESA = lp.C_INT_CODIGO_EMPRESA
                ORDER BY lp.C_DAT_FECHA_MOVIMIENTO
            """)

            registros = []
            for row in cur.fetchall():
                nro_letra, fecha_mov, total, moneda_cod, proveedor, empresa, obs = row
                # Calcular monto en PEN
                if moneda_cod.upper() == 'PEN':
                    monto_pen = float(total or 0)
                else:
                    tc = get_tc(self._params_base, moneda_cod)
                    monto_pen = round(float(total or 0) * tc, 2)

                fecha_str = fecha_mov.strftime("%Y-%m-%d") if fecha_mov else ""
                nota = nro_letra.strip() if nro_letra else ""
                obs_clean = obs.strip()
                if obs_clean:
                    nota += f" | {obs_clean}"

                registros.append({
                    "fecha_vcto": fecha_str,
                    "monto_pen":  round(monto_pen, 2),
                    "empresa":    empresa.strip(),
                    "proveedor":  proveedor.strip(),
                    "nota":       nota,
                    "_nro_letra": nro_letra.strip() if nro_letra else "",
                })

            conn.close()
            self.done.emit(registros)
        except Exception as ex:
            self.error.emit(str(ex))


# ── Tab: Base de Datos ─────────────────────────────────────────────────────────
_FILTER_STYLE = (
    "QLineEdit { border: 1px solid #aaa; border-radius: 3px; padding: 2px 4px; "
    "background: #fffbe6; font-style: italic; }"
)

def _tbl_cell(tbl, row, col):
    item = tbl.item(row, col)
    return item.text().strip() if item else ""

def _apply_filter(tbl, col_filters):
    """Muestra/oculta filas según los filtros [(col_idx, texto), ...]."""
    visible = 0
    for i in range(tbl.rowCount()):
        match = all(
            not txt or txt in _tbl_cell(tbl, i, col).lower()
            for col, txt in col_filters
        )
        tbl.setRowHidden(i, not match)
        if match:
            visible += 1
    return visible


class TabBase(QWidget):
    base_changed = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self._base = load_base()
        self._build()
        self._load_ui()

    def _build(self):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        container = QWidget()
        v = QVBoxLayout(container)
        v.setSpacing(8)
        scroll.setWidget(container)
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.addWidget(scroll)

        # ── Proveedores ───────────────────────────────────────────────────────
        grp_prov = QGroupBox("Proveedores")
        vp = QVBoxLayout(grp_prov)

        fp = QHBoxLayout()
        fp.addWidget(QLabel("🔍"))
        self.flt_prov_nom = QLineEdit()
        self.flt_prov_nom.setPlaceholderText("Filtrar por nombre…")
        self.flt_prov_nom.setStyleSheet(_FILTER_STYLE)
        self.flt_prov_nom.textChanged.connect(self._filter_prov)
        self.flt_prov_ruc = QLineEdit()
        self.flt_prov_ruc.setPlaceholderText("Filtrar por RUC…")
        self.flt_prov_ruc.setStyleSheet(_FILTER_STYLE)
        self.flt_prov_ruc.setMaximumWidth(160)
        self.flt_prov_ruc.textChanged.connect(self._filter_prov)
        fp.addWidget(self.flt_prov_nom)
        fp.addWidget(self.flt_prov_ruc)
        vp.addLayout(fp)

        self.tbl_prov = QTableWidget(0, 2)
        self.tbl_prov.setHorizontalHeaderLabels(["Nombre del Proveedor", "RUC"])
        self.tbl_prov.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.tbl_prov.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.tbl_prov.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tbl_prov.setMinimumHeight(200)
        vp.addWidget(self.tbl_prov)

        hp = QHBoxLayout()
        self.lbl_prov_count = QLabel("0 proveedores")
        btn_add_prov = QPushButton("+ Agregar fila")
        btn_add_prov.clicked.connect(self._add_prov)
        btn_del_prov = QPushButton("🗑 Eliminar seleccionado")
        btn_del_prov.clicked.connect(self._del_prov)
        hp.addWidget(self.lbl_prov_count)
        hp.addStretch()
        hp.addWidget(btn_add_prov)
        hp.addWidget(btn_del_prov)
        vp.addLayout(hp)
        v.addWidget(grp_prov)

        # ── Empresas Emisoras ─────────────────────────────────────────────────
        grp_emp = QGroupBox("Empresas Emisoras")
        ve = QVBoxLayout(grp_emp)

        fe = QHBoxLayout()
        fe.addWidget(QLabel("🔍"))
        self.flt_emp = QLineEdit()
        self.flt_emp.setPlaceholderText("Filtrar por nombre…")
        self.flt_emp.setStyleSheet(_FILTER_STYLE)
        self.flt_emp.textChanged.connect(self._filter_emp)
        fe.addWidget(self.flt_emp)
        ve.addLayout(fe)

        self.tbl_emp = QTableWidget(0, 1)
        self.tbl_emp.setHorizontalHeaderLabels(["Nombre Empresa"])
        self.tbl_emp.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.tbl_emp.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tbl_emp.setMaximumHeight(160)
        ve.addWidget(self.tbl_emp)

        he = QHBoxLayout()
        self.lbl_emp_count = QLabel("0 empresas")
        btn_add_emp = QPushButton("+ Agregar fila")
        btn_add_emp.clicked.connect(self._add_emp)
        btn_del_emp = QPushButton("🗑 Eliminar seleccionada")
        btn_del_emp.clicked.connect(self._del_emp)
        he.addWidget(self.lbl_emp_count)
        he.addStretch()
        he.addWidget(btn_add_emp)
        he.addWidget(btn_del_emp)
        ve.addLayout(he)
        v.addWidget(grp_emp)

        # ── Temporadas ────────────────────────────────────────────────────────
        grp_temp = QGroupBox("Temporadas")
        vt = QVBoxLayout(grp_temp)

        ft = QHBoxLayout()
        ft.addWidget(QLabel("🔍"))
        self.flt_temp = QLineEdit()
        self.flt_temp.setPlaceholderText("Filtrar por prefijo…")
        self.flt_temp.setStyleSheet(_FILTER_STYLE)
        self.flt_temp.textChanged.connect(self._filter_temp)
        ft.addWidget(self.flt_temp)
        vt.addLayout(ft)

        self.tbl_temp = QTableWidget(0, 1)
        self.tbl_temp.setHorizontalHeaderLabels(["Prefijo Temporada"])
        self.tbl_temp.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.tbl_temp.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tbl_temp.setMaximumHeight(140)
        vt.addWidget(self.tbl_temp)

        ht = QHBoxLayout()
        btn_add_temp = QPushButton("+ Agregar fila")
        btn_add_temp.clicked.connect(self._add_temp)
        btn_del_temp = QPushButton("🗑 Eliminar seleccionada")
        btn_del_temp.clicked.connect(self._del_temp)
        ht.addStretch()
        ht.addWidget(btn_add_temp)
        ht.addWidget(btn_del_temp)
        vt.addLayout(ht)
        v.addWidget(grp_temp)

        # ── Importar desde DYSPOW ─────────────────────────────────────────────
        grp_dyspow = QGroupBox("Importar desde DYSPOW (solo lectura)")
        hd = QHBoxLayout(grp_dyspow)
        self.btn_imp_prov = QPushButton("🔄 Importar Proveedores")
        self.btn_imp_emp  = QPushButton("🔄 Importar Empresas")
        self.lbl_dyspow   = QLabel("")
        self.btn_imp_prov.clicked.connect(self._importar_proveedores)
        self.btn_imp_emp.clicked.connect(self._importar_empresas)
        hd.addWidget(self.btn_imp_prov)
        hd.addWidget(self.btn_imp_emp)
        hd.addWidget(self.lbl_dyspow)
        hd.addStretch()
        v.addWidget(grp_dyspow)

        btn_save = QPushButton("💾 Guardar Base de Datos")
        btn_save.setFixedHeight(36)
        btn_save.setStyleSheet("font-weight:bold")
        btn_save.clicked.connect(self._guardar)
        v.addWidget(btn_save)
        v.addStretch()

    # ── Filtros ───────────────────────────────────────────────────────────────
    def _filter_prov(self):
        vis = _apply_filter(self.tbl_prov, [
            (0, self.flt_prov_nom.text().lower()),
            (1, self.flt_prov_ruc.text().lower()),
        ])
        self.lbl_prov_count.setText(f"{vis} de {self.tbl_prov.rowCount()} proveedores")

    def _filter_emp(self):
        vis = _apply_filter(self.tbl_emp, [(0, self.flt_emp.text().lower())])
        self.lbl_emp_count.setText(f"{vis} de {self.tbl_emp.rowCount()} empresas")

    def _filter_temp(self):
        _apply_filter(self.tbl_temp, [(0, self.flt_temp.text().lower())])

    # ── Agregar filas ─────────────────────────────────────────────────────────
    def _add_prov(self):
        # Limpiar filtro para que la nueva fila sea visible
        self.flt_prov_nom.clear(); self.flt_prov_ruc.clear()
        i = self.tbl_prov.rowCount()
        self.tbl_prov.insertRow(i)
        self.tbl_prov.scrollToBottom()
        self.tbl_prov.setCurrentCell(i, 0)
        self.tbl_prov.editItem(self.tbl_prov.item(i, 0) or QTableWidgetItem(""))
        self._filter_prov()

    def _add_emp(self):
        self.flt_emp.clear()
        i = self.tbl_emp.rowCount()
        self.tbl_emp.insertRow(i)
        self.tbl_emp.scrollToBottom()
        self.tbl_emp.setCurrentCell(i, 0)
        self._filter_emp()

    def _add_temp(self):
        self.flt_temp.clear()
        i = self.tbl_temp.rowCount()
        self.tbl_temp.insertRow(i)
        self.tbl_temp.scrollToBottom()
        self.tbl_temp.setCurrentCell(i, 0)

    # ── Eliminar filas ────────────────────────────────────────────────────────
    def _del_prov(self):
        row = self.tbl_prov.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Sin selección", "Seleccione un proveedor.")
            return
        nom = _tbl_cell(self.tbl_prov, row, 0) or "(vacío)"
        if QMessageBox.question(self, "Confirmar", f"¿Eliminar '{nom}'?",
                                QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            self.tbl_prov.removeRow(row)
            self._filter_prov()

    def _del_emp(self):
        row = self.tbl_emp.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Sin selección", "Seleccione una empresa.")
            return
        nom = _tbl_cell(self.tbl_emp, row, 0) or "(vacío)"
        if QMessageBox.question(self, "Confirmar", f"¿Eliminar '{nom}'?",
                                QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            self.tbl_emp.removeRow(row)
            self._filter_emp()

    def _del_temp(self):
        row = self.tbl_temp.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Sin selección", "Seleccione una temporada.")
            return
        nom = _tbl_cell(self.tbl_temp, row, 0) or "(vacío)"
        if QMessageBox.question(self, "Confirmar", f"¿Eliminar '{nom}'?",
                                QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            self.tbl_temp.removeRow(row)

    # ── Cargar / Guardar ──────────────────────────────────────────────────────
    def _load_ui(self):
        self.tbl_prov.setRowCount(0)
        for p in self._base.get("proveedores", []):
            i = self.tbl_prov.rowCount()
            self.tbl_prov.insertRow(i)
            self.tbl_prov.setItem(i, 0, QTableWidgetItem(p.get("nombre", "")))
            self.tbl_prov.setItem(i, 1, QTableWidgetItem(p.get("ruc", "")))
        self._filter_prov()

        self.tbl_emp.setRowCount(0)
        for e in self._base.get("empresas", []):
            i = self.tbl_emp.rowCount()
            self.tbl_emp.insertRow(i)
            self.tbl_emp.setItem(i, 0, QTableWidgetItem(e))
        self._filter_emp()

        self.tbl_temp.setRowCount(0)
        for t in self._base.get("temporadas", []):
            i = self.tbl_temp.rowCount()
            self.tbl_temp.insertRow(i)
            self.tbl_temp.setItem(i, 0, QTableWidgetItem(t))

    def _guardar(self):
        provs = []
        for i in range(self.tbl_prov.rowCount()):
            nom = _tbl_cell(self.tbl_prov, i, 0)
            ruc = _tbl_cell(self.tbl_prov, i, 1)
            if nom:
                provs.append({"nombre": nom, "ruc": ruc})
        self._base["proveedores"] = provs

        emps = []
        for i in range(self.tbl_emp.rowCount()):
            e = _tbl_cell(self.tbl_emp, i, 0)
            if e:
                emps.append(e)
        self._base["empresas"] = emps

        temps = []
        for i in range(self.tbl_temp.rowCount()):
            t = _tbl_cell(self.tbl_temp, i, 0)
            if t:
                temps.append(t)
        self._base["temporadas"] = temps

        save_base(self._base)
        QMessageBox.information(self, "Guardado", "Base de datos guardada.")
        self.base_changed.emit()

    # ── Importar desde DYSPOW ─────────────────────────────────────────────────
    def _get_dyspow_cfg(self):
        cfg = load_dyspow_cfg()
        dlg = DyspowConfigDialog(self)
        dlg.txt_server.setText(cfg.get("server", ""))
        dlg.txt_database.setText(cfg.get("database", ""))
        dlg.txt_user.setText(cfg.get("user", ""))
        dlg.txt_password.setText(cfg.get("password", ""))
        if dlg.exec() != QDialog.Accepted:
            return None
        cfg = dlg.get_config()
        save_dyspow_cfg(cfg)
        return cfg

    def _importar_proveedores(self):
        cfg = self._get_dyspow_cfg()
        if not cfg:
            return
        self.btn_imp_prov.setEnabled(False)
        self.btn_imp_emp.setEnabled(False)
        self.lbl_dyspow.setText("Conectando…")
        self._worker_base = WorkerDyspowProveedores(cfg, self)
        self._worker_base.done.connect(self._on_dyspow_prov_done)
        self._worker_base.error.connect(self._on_dyspow_error)
        self._worker_base.start()

    def _importar_empresas(self):
        cfg = self._get_dyspow_cfg()
        if not cfg:
            return
        self.btn_imp_prov.setEnabled(False)
        self.btn_imp_emp.setEnabled(False)
        self.lbl_dyspow.setText("Conectando…")
        self._worker_emp = WorkerDyspowProveedores(cfg, self)
        self._worker_emp.done.connect(self._on_dyspow_emp_done)
        self._worker_emp.error.connect(self._on_dyspow_error)
        self._worker_emp.start()

    def _on_dyspow_prov_done(self, provs_nuevos, _emps):
        self.btn_imp_prov.setEnabled(True)
        self.btn_imp_emp.setEnabled(True)

        rucs_existentes = set()
        nombres_existentes = set()
        for i in range(self.tbl_prov.rowCount()):
            rucs_existentes.add(_tbl_cell(self.tbl_prov, i, 1))
            nombres_existentes.add(_tbl_cell(self.tbl_prov, i, 0).upper())

        nuevos = [
            p for p in provs_nuevos
            if not (p["ruc"] and p["ruc"] in rucs_existentes)
            and not (not p["ruc"] and p["nombre"].upper() in nombres_existentes)
        ]

        if not nuevos:
            self.lbl_dyspow.setText(f"Sin novedades ({len(provs_nuevos)} ya existían)")
            return

        self.tbl_prov.setSortingEnabled(False)
        for p in nuevos:
            i = self.tbl_prov.rowCount()
            self.tbl_prov.insertRow(i)
            self.tbl_prov.setItem(i, 0, QTableWidgetItem(p["nombre"]))
            self.tbl_prov.setItem(i, 1, QTableWidgetItem(p["ruc"]))
        self._filter_prov()
        self.lbl_dyspow.setText(f"✓ {len(nuevos)} proveedores importados")

    def _on_dyspow_emp_done(self, _provs, emps_nuevas):
        self.btn_imp_prov.setEnabled(True)
        self.btn_imp_emp.setEnabled(True)

        existentes = {_tbl_cell(self.tbl_emp, i, 0).upper()
                      for i in range(self.tbl_emp.rowCount())}
        nuevas = [e for e in emps_nuevas if e.upper() not in existentes]

        if not nuevas:
            self.lbl_dyspow.setText(f"Sin novedades ({len(emps_nuevas)} ya existían)")
            return

        for emp in nuevas:
            i = self.tbl_emp.rowCount()
            self.tbl_emp.insertRow(i)
            self.tbl_emp.setItem(i, 0, QTableWidgetItem(emp))
        self._filter_emp()
        self.lbl_dyspow.setText(f"✓ {len(nuevas)} empresas importadas")

    def _on_dyspow_error(self, msg):
        self.btn_imp_prov.setEnabled(True)
        self.btn_imp_emp.setEnabled(True)
        self.lbl_dyspow.setText("Error de conexión")
        QMessageBox.critical(self, "Error DYSPOW", f"No se pudo conectar:\n{msg}")

    def get_base(self):
        return self._base


# ── Tab: Datos Externos ────────────────────────────────────────────────────────
class TabExternos(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._externos = load_ext()
        self._build()
        self._refresh()

    def _build(self):
        v = QVBoxLayout(self)
        v.addWidget(QLabel("⚠  Letras emitidas fuera del Generador. Se usan para calcular el compromiso diario."))

        # Fila de filtros (alineada con columnas de la tabla)
        filt_row = QHBoxLayout()
        filt_row.addWidget(QLabel("🔍"))
        self.flt_fecha  = QLineEdit(); self.flt_fecha.setPlaceholderText("Fecha…");    self.flt_fecha.setMaximumWidth(100)
        self.flt_monto  = QLineEdit(); self.flt_monto.setPlaceholderText("Monto…");    self.flt_monto.setMaximumWidth(100)
        self.flt_emp    = QLineEdit(); self.flt_emp.setPlaceholderText("Empresa…");    self.flt_emp.setMaximumWidth(160)
        self.flt_prov   = QLineEdit(); self.flt_prov.setPlaceholderText("Proveedor…")
        self.flt_nota   = QLineEdit(); self.flt_nota.setPlaceholderText("Nota…")
        for w in (self.flt_fecha, self.flt_monto, self.flt_emp, self.flt_prov, self.flt_nota):
            w.setStyleSheet(_FILTER_STYLE)
            filt_row.addWidget(w)
        btn_clear_flt = QPushButton("✕")
        btn_clear_flt.setMaximumWidth(28)
        btn_clear_flt.setToolTip("Limpiar filtros")
        btn_clear_flt.clicked.connect(self._clear_filters)
        filt_row.addWidget(btn_clear_flt)
        v.addLayout(filt_row)

        for w in (self.flt_fecha, self.flt_monto, self.flt_emp, self.flt_prov, self.flt_nota):
            w.textChanged.connect(self._filter_ext)

        self.tbl = QTableWidget(0, 5)
        self.tbl.setHorizontalHeaderLabels(["Fecha Vcto.", "Monto PEN", "Empresa", "Proveedor", "Nota"])
        self.tbl.horizontalHeader().setSectionResizeMode(4, QHeaderView.Stretch)
        for c in (0, 1, 2, 3):
            self.tbl.horizontalHeader().setSectionResizeMode(c, QHeaderView.ResizeToContents)
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
        v.addWidget(self.tbl)

        h = QHBoxLayout()
        self.lbl_ext_count = QLabel("0 registros")
        btn_add = QPushButton("+ Agregar fila")
        btn_add.clicked.connect(self._agregar_fila)
        btn_del = QPushButton("🗑 Eliminar seleccionada")
        btn_del.clicked.connect(self._eliminar_fila)
        self.btn_dyspow = QPushButton("🔄 Importar Letras de DYSPOW")
        self.btn_dyspow.clicked.connect(self._importar_letras_dyspow)
        self.lbl_dyspow = QLabel("")
        btn_save = QPushButton("💾 Guardar")
        btn_save.clicked.connect(self._guardar)
        h.addWidget(self.lbl_ext_count)
        h.addWidget(btn_add)
        h.addWidget(btn_del)
        h.addWidget(self.btn_dyspow)
        h.addWidget(self.lbl_dyspow)
        h.addStretch()
        h.addWidget(btn_save)
        v.addLayout(h)

    def _refresh(self):
        self.tbl.setRowCount(0)
        for e in self._externos:
            i = self.tbl.rowCount()
            self.tbl.insertRow(i)
            self.tbl.setItem(i, 0, QTableWidgetItem(e.get("fecha_vcto", "")))
            self.tbl.setItem(i, 1, QTableWidgetItem(str(e.get("monto_pen", 0))))
            self.tbl.setItem(i, 2, QTableWidgetItem(e.get("empresa", "")))
            self.tbl.setItem(i, 3, QTableWidgetItem(e.get("proveedor", "")))
            self.tbl.setItem(i, 4, QTableWidgetItem(e.get("nota", "")))
        self._filter_ext()

    def _filter_ext(self):
        vis = _apply_filter(self.tbl, [
            (0, self.flt_fecha.text().lower()),
            (1, self.flt_monto.text().lower()),
            (2, self.flt_emp.text().lower()),
            (3, self.flt_prov.text().lower()),
            (4, self.flt_nota.text().lower()),
        ])
        self.lbl_ext_count.setText(f"{vis} de {self.tbl.rowCount()} registros")

    def _clear_filters(self):
        for w in (self.flt_fecha, self.flt_monto, self.flt_emp, self.flt_prov, self.flt_nota):
            w.clear()

    def _agregar_fila(self):
        self._clear_filters()
        i = self.tbl.rowCount()
        self.tbl.insertRow(i)
        self.tbl.scrollToBottom()
        self.tbl.setCurrentCell(i, 0)
        self._filter_ext()

    def _eliminar_fila(self):
        row = self.tbl.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Sin selección", "Seleccione una fila.")
            return
        if QMessageBox.question(self, "Confirmar", "¿Eliminar esta fila?",
                                QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            self.tbl.removeRow(row)
            self._filter_ext()

    def _guardar(self):
        ext = []
        for i in range(self.tbl.rowCount()):
            fv   = (self.tbl.item(i,0).text() if self.tbl.item(i,0) else "").strip()
            try:
                monto = float((self.tbl.item(i,1).text() if self.tbl.item(i,1) else "0").replace(",",".") or 0)
            except Exception:
                monto = 0
            emp  = (self.tbl.item(i,2).text() if self.tbl.item(i,2) else "").strip()
            prov = (self.tbl.item(i,3).text() if self.tbl.item(i,3) else "").strip()
            nota = (self.tbl.item(i,4).text() if self.tbl.item(i,4) else "").strip()
            if fv:
                ext.append({"fecha_vcto": fv, "monto_pen": monto, "empresa": emp, "proveedor": prov, "nota": nota})
        self._externos = ext
        save_ext(ext)
        QMessageBox.information(self, "Guardado", "Datos externos guardados.")

    def _importar_letras_dyspow(self):
        # Mostrar diálogo de configuración
        cfg = load_dyspow_cfg()
        dlg = DyspowConfigDialog(self)
        dlg.txt_server.setText(cfg.get("server", ""))
        dlg.txt_database.setText(cfg.get("database", ""))
        dlg.txt_user.setText(cfg.get("user", ""))
        dlg.txt_password.setText(cfg.get("password", ""))
        if dlg.exec() != QDialog.Accepted:
            return
        cfg = dlg.get_config()
        save_dyspow_cfg(cfg)

        self.btn_dyspow.setEnabled(False)
        self.lbl_dyspow.setText("Conectando…")
        params_base = load_base()
        self._worker_letras = WorkerDyspowLetras(cfg, params_base, self)
        self._worker_letras.done.connect(self._on_dyspow_letras_done)
        self._worker_letras.error.connect(self._on_dyspow_letras_error)
        self._worker_letras.start()

    def _on_dyspow_letras_done(self, registros: list):
        self.btn_dyspow.setEnabled(True)

        # Determinar números de letra ya registrados (evitar duplicados)
        notas_existentes = set()
        for e in self._externos:
            nota = e.get("nota", "")
            # extraer el número antes del " | "
            nro = nota.split(" | ")[0].strip()
            if nro:
                notas_existentes.add(nro)

        nuevos = []
        for r in registros:
            nro = r.get("_nro_letra", "")
            if nro and nro in notas_existentes:
                continue
            nuevos.append({
                "fecha_vcto": r["fecha_vcto"],
                "monto_pen":  r["monto_pen"],
                "empresa":    r["empresa"],
                "proveedor":  r["proveedor"],
                "nota":       r["nota"],
            })

        if not nuevos:
            self.lbl_dyspow.setText(f"Sin novedades ({len(registros)} ya existían)")
            return

        # Agregar al final de la tabla y de _externos
        for reg in nuevos:
            i = self.tbl.rowCount()
            self.tbl.insertRow(i)
            self.tbl.setItem(i, 0, QTableWidgetItem(reg["fecha_vcto"]))
            self.tbl.setItem(i, 1, QTableWidgetItem(str(reg["monto_pen"])))
            self.tbl.setItem(i, 2, QTableWidgetItem(reg["empresa"]))
            self.tbl.setItem(i, 3, QTableWidgetItem(reg["proveedor"]))
            self.tbl.setItem(i, 4, QTableWidgetItem(reg["nota"]))
            self._externos.append(reg)

        save_ext(self._externos)
        self.lbl_dyspow.setText(f"✓ {len(nuevos)} letras importadas")

    def _on_dyspow_letras_error(self, msg):
        self.btn_dyspow.setEnabled(True)
        self.lbl_dyspow.setText("Error de conexión")
        QMessageBox.critical(self, "Error DYSPOW", f"No se pudo conectar:\n{msg}")

    def get_externos(self):
        return self._externos


# ── Ventana Principal ──────────────────────────────────────────────────────────
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Generador de Letras de Cambio")
        self.setMinimumSize(1000, 720)

        self._historico = load_hist()
        self._build()

    def _build(self):
        tabs = QTabWidget()
        self.setCentralWidget(tabs)

        # Tab Base
        self.tab_base = TabBase()
        self.tab_base.base_changed.connect(self._on_base_changed)
        tabs.addTab(self.tab_base, "📋 Base de Datos")

        # Tab Parámetros
        self.tab_params = TabParametros()
        self.tab_params.params_changed.connect(self._on_params_changed)
        tabs.addTab(self.tab_params, "⚙ Parámetros")

        # Tab Generador
        self.tab_gen = TabGenerador(self._get_data)
        self.tab_gen.cronograma_generado.connect(self._on_cronograma_generado)
        tabs.addTab(self.tab_gen, "⚡ Generador")

        # Tab Histórico
        self.tab_hist = TabHistorico()
        self.tab_hist.refresh(self._historico)
        tabs.addTab(self.tab_hist, "📜 Histórico")

        # Tab Externos
        self.tab_ext = TabExternos()
        tabs.addTab(self.tab_ext, "🔗 Datos Externos")

        # Inicializar combos del generador
        self.tab_gen.refresh_combos(self.tab_base.get_base())

        # Ir al generador por defecto
        tabs.setCurrentIndex(2)

    def _get_data(self):
        return (
            self.tab_params.get_params(),
            self.tab_base.get_base(),
            self._historico,
            self.tab_ext.get_externos(),
        )

    def _on_base_changed(self):
        self.tab_gen.refresh_combos(self.tab_base.get_base())

    def _on_params_changed(self):
        pass  # params se releen en tiempo real desde get_params()

    def _on_cronograma_generado(self, letras: list):
        ahora = datetime.now().isoformat(timespec="seconds")
        id_op = gen_id(self._historico)
        for lt in letras:
            self._historico.append({
                "id_operacion":  id_op,
                "fecha_registro": ahora,
                "empresa":       lt["empresa"],
                "proveedor":     lt["proveedor"],
                "ruc":           lt.get("ruc", ""),
                "concepto":      lt["concepto"],
                "observac":      lt.get("observac", ""),
                "temporada":     lt["temporada"],
                "n_letra":       lt["n"],
                "moneda":        lt["moneda"],
                "fecha_emi":     lt["fecha_emi"],
                "fecha_vcto":    lt["fecha_vcto"],
                "monto":         lt["monto"],
            })
        save_hist(self._historico)
        self.tab_hist.refresh(self._historico)


# ── Entry point ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    # Verificar dateutil
    try:
        from dateutil.relativedelta import relativedelta
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "python-dateutil", "-q"])

    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    win = MainWindow()
    win.show()
    sys.exit(app.exec())
